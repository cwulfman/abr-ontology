#!/usr/bin/env python3
"""
Generate an enhanced spreadsheet for Rock the Vote administrators.
Creates an Excel file with:
- Grouped/hierarchical columns
- Color coding
- Summary statistics
- Professional formatting for non-technical audience
"""

import re
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Installing it now...")
    import subprocess
    subprocess.check_call(['pip3', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True

# Define the information type hierarchy
INFO_HIERARCHY = {
    'Election Context': [
        'ElectionDate',
        'ElectionType',
        'ElectionParty',
        'ElectionRangeRequest',
        'SpecificElectionIdentifier',
        'ElectionContext',
    ],
    'Eligibility Criteria': [
        'CitizenshipStatus',
        'AgeRequirement',
        'ResidencyRequirement',
        'RegistrationLocation',
        'MilitaryOrOverseasStatus',
        'DisabilityStatus',
        'StudentStatus',
        'AddressConfidentialityProgramParticipation',
    ],
    'Reason for Absentee Request': [
        'OutOfCountyTravelReason',
        'IllnessOrDisabilityReason',
        'CaregiverForDisabledPersonReason',
        'WorkConflictReason',
        'ReligiousConflictReason',
        'ElectionOfficialDutiesReason',
        'IncarcerationReason',
        'NursingHomeReason',
        'ExpectedChildbirthReason',
    ],
    'Personal Info: Identifiers': [
        'PersonName',
        'FirstName',
        'LastName',
        'MiddleName',
        'NameSuffix',
        'FormerName',
        'ResidentialAddress',
        'MailingAddress',
        'PhoneNumber',
        'EmailAddress',
    ],
    'Personal Info: ID Numbers': [
        'DriverLicenseNumber',
        'StateIDNumber',
        'Last4SSN',
        'VoterPIN',
        'TribalID',
        'PassportOrMilitaryID',
        'IdentificationNumbers',
    ],
    'Personal Info: Demographics': [
        'DateOfBirth',
        'PlaceOfBirth',
        'Gender',
    ],
    'Ballot Delivery': [
        'BallotMailingAddress',
        'BallotEmailAddress',
        'BallotFaxNumber',
        'BallotDeliveryAgent',
        'AgentName',
        'AgentRelationship',
        'AgentContactInformation',
        'BallotDeliveryMethod',
        'TemporaryAbsenceAddress',
    ],
    'Party Affiliation': [
        'PartyAffiliation',
        'PrimaryBallotChoice',
        'BallotStyleRequest',
    ],
    'Signature & Attestation': [
        'VoterSignature',
        'SignatureDate',
        'WitnessInformation',
        'WitnessName',
        'WitnessAddress',
        'NotaryInformation',
        'NotaryName',
        'NotaryCommissionExpiration',
        'AssistantInformation',
        'AttesterInformation',
    ],
    'Required Documentation': [
        'CopyOfPhotoID',
        'ProofOfResidence',
        'IDNumbersProvided',
        'StatementNoIDAvailable',
    ],
    'Required Accompanying Docs': [
        'RequiredPhotoID',
        'RequiredProofOfResidence',
        'RequiredWitnessSignature',
        'RequiredNotary',
        'RequiredAgentCertification',
    ],
}

# Human-readable labels for information types
INFO_LABELS = {
    'ElectionDate': 'Election Date',
    'ElectionType': 'Election Type',
    'ElectionParty': 'Election Party',
    'ElectionRangeRequest': 'Election Range',
    'SpecificElectionIdentifier': 'Specific Election ID',
    'ElectionContext': 'General Election Info',
    'CitizenshipStatus': 'Citizenship',
    'AgeRequirement': 'Age',
    'ResidencyRequirement': 'Residency',
    'RegistrationLocation': 'Registration Location',
    'MilitaryOrOverseasStatus': 'Military/Overseas',
    'DisabilityStatus': 'Disability',
    'StudentStatus': 'Student',
    'AddressConfidentialityProgramParticipation': 'Address Confidentiality',
    'OutOfCountyTravelReason': 'Out of County Travel',
    'IllnessOrDisabilityReason': 'Illness/Disability',
    'CaregiverForDisabledPersonReason': 'Caregiver',
    'WorkConflictReason': 'Work Conflict',
    'ReligiousConflictReason': 'Religious Conflict',
    'ElectionOfficialDutiesReason': 'Election Official Duties',
    'IncarcerationReason': 'Incarceration',
    'NursingHomeReason': 'Nursing Home',
    'ExpectedChildbirthReason': 'Expected Childbirth',
    'PersonName': 'Name',
    'FirstName': 'First Name',
    'LastName': 'Last Name',
    'MiddleName': 'Middle Name',
    'NameSuffix': 'Name Suffix',
    'FormerName': 'Former Name',
    'ResidentialAddress': 'Residential Address',
    'MailingAddress': 'Mailing Address',
    'PhoneNumber': 'Phone',
    'EmailAddress': 'Email',
    'DriverLicenseNumber': "Driver's License #",
    'StateIDNumber': 'State ID #',
    'Last4SSN': 'Last 4 SSN',
    'VoterPIN': 'Voter PIN',
    'TribalID': 'Tribal ID',
    'PassportOrMilitaryID': 'Passport/Military ID',
    'IdentificationNumbers': 'ID Numbers (General)',
    'DateOfBirth': 'Date of Birth',
    'PlaceOfBirth': 'Place of Birth',
    'Gender': 'Gender',
    'BallotMailingAddress': 'Mailing Address',
    'BallotEmailAddress': 'Email Address',
    'BallotFaxNumber': 'Fax Number',
    'BallotDeliveryAgent': 'Delivery Agent',
    'AgentName': 'Agent Name',
    'AgentRelationship': 'Agent Relationship',
    'AgentContactInformation': 'Agent Contact',
    'BallotDeliveryMethod': 'Delivery Method',
    'TemporaryAbsenceAddress': 'Temporary Address',
    'PartyAffiliation': 'Party Affiliation',
    'PrimaryBallotChoice': 'Primary Ballot',
    'BallotStyleRequest': 'Ballot Style',
    'VoterSignature': 'Voter Signature',
    'SignatureDate': 'Signature Date',
    'WitnessInformation': 'Witness Info',
    'WitnessName': 'Witness Name',
    'WitnessAddress': 'Witness Address',
    'NotaryInformation': 'Notary Info',
    'NotaryName': 'Notary Name',
    'NotaryCommissionExpiration': 'Notary Expiration',
    'AssistantInformation': 'Assistant Info',
    'AttesterInformation': 'Attester Info',
    'CopyOfPhotoID': 'Copy of Photo ID',
    'ProofOfResidence': 'Proof of Residence',
    'IDNumbersProvided': 'ID Numbers',
    'StatementNoIDAvailable': 'Statement: No ID',
    'RequiredPhotoID': 'Photo ID',
    'RequiredProofOfResidence': 'Proof of Residence',
    'RequiredWitnessSignature': 'Witness Signature',
    'RequiredNotary': 'Notary',
    'RequiredAgentCertification': 'Agent/Assistant Certification',
}

ALL_STATES = [
    'AK', 'AL', 'AR', 'AZ', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA',
    'HI', 'IA', 'ID', 'IL', 'IN', 'KS', 'KY', 'LA', 'MA', 'MD',
    'ME', 'MI', 'MN', 'MO', 'MS', 'MT', 'NC', 'ND', 'NE', 'NH',
    'NJ', 'NM', 'NV', 'NY', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC',
    'SD', 'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WI', 'WV', 'WY'
]

STATE_NAMES = {
    'AK': 'Alaska', 'AL': 'Alabama', 'AR': 'Arkansas', 'AZ': 'Arizona',
    'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
    'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii', 'IA': 'Iowa',
    'ID': 'Idaho', 'IL': 'Illinois', 'IN': 'Indiana', 'KS': 'Kansas',
    'KY': 'Kentucky', 'LA': 'Louisiana', 'MA': 'Massachusetts', 'MD': 'Maryland',
    'ME': 'Maine', 'MI': 'Michigan', 'MN': 'Minnesota', 'MO': 'Missouri',
    'MS': 'Mississippi', 'MT': 'Montana', 'NC': 'North Carolina', 'ND': 'North Dakota',
    'NE': 'Nebraska', 'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico',
    'NV': 'Nevada', 'NY': 'New York', 'OH': 'Ohio', 'OK': 'Oklahoma',
    'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
    'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah',
    'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington', 'WI': 'Wisconsin',
    'WV': 'West Virginia', 'WY': 'Wyoming'
}

def parse_ontology(ontology_file):
    """Parse the ontology and extract state requirements."""
    state_info = defaultdict(set)

    with open(ontology_file, 'r') as f:
        content = f.read()

    state_forms = re.findall(r'oset:([A-Z]{2})_Form', content)
    state_forms = list(set(state_forms))

    for state in state_forms:
        pattern = rf'oset:{state}_Form\s+(?:a\s+owl:Class\s*;)?\s*rdfs:subClassOf.*?(?=(?:oset:[A-Z]{{2}}_Form|#{{50,}}|\Z))'
        matches = re.findall(pattern, content, re.DOTALL)

        for match in matches:
            info_types = re.findall(r'owl:someValuesFrom\s+oset:(\w+)', match)
            state_info[state].update(info_types)

    return state_info

def create_excel_spreadsheet(state_info, output_file):
    """Create a professionally formatted Excel spreadsheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "State ABR Requirements"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    na_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    bold_font = Font(bold=True)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Create header rows
    row1 = ["State", "Code"]
    row2 = ["", ""]

    col_idx = 3
    category_ranges = {}

    for category, info_types in INFO_HIERARCHY.items():
        start_col = col_idx
        for info_type in info_types:
            row1.append(category)
            row2.append(INFO_LABELS.get(info_type, info_type))
            col_idx += 1
        end_col = col_idx - 1
        category_ranges[category] = (start_col, end_col)

    # Write headers
    ws.append(row1)
    ws.append(row2)

    # Format header rows
    for col in range(1, col_idx):
        cell1 = ws.cell(row=1, column=col)
        cell2 = ws.cell(row=2, column=col)

        if col <= 2:
            cell1.fill = header_fill
            cell1.font = header_font
            cell2.fill = header_fill
            cell2.font = header_font
        else:
            cell1.fill = header_fill
            cell1.font = header_font
            cell2.fill = subheader_fill
            cell2.font = subheader_font

        cell1.alignment = center_align
        cell2.alignment = center_align
        cell1.border = border
        cell2.border = border

    # Merge category headers
    ws.merge_cells(range_string='A1:A2')
    ws.merge_cells(range_string='B1:B2')

    for category, (start_col, end_col) in category_ranges.items():
        if start_col < end_col:
            ws.merge_cells(
                start_row=1,
                start_column=start_col,
                end_row=1,
                end_column=end_col
            )

    # Add data rows (only for states with forms)
    row_num = 3
    for state_code in sorted(state_info.keys()):
        row_data = [STATE_NAMES[state_code], state_code]

        # Information types
        for category, info_types in INFO_HIERARCHY.items():
            for info_type in info_types:
                row_data.append("✓" if info_type in state_info[state_code] else "")

        ws.append(row_data)

        # Format data cells
        for col in range(1, col_idx):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = center_align

            if col > 2:  # Information type columns
                if cell.value == "✓":
                    cell.fill = yes_fill
                    cell.font = Font(color="006100", bold=True, size=12)

        row_num += 1

    # Add summary row
    summary_row = ["TOTAL", f"{len(state_info)} states"]
    for category, info_types in INFO_HIERARCHY.items():
        for info_type in info_types:
            count = sum(1 for state in state_info.keys() if info_type in state_info[state])
            summary_row.append(f"{count}")

    ws.append(summary_row)

    # Format summary row
    for col in range(1, col_idx):
        cell = ws.cell(row=row_num, column=col)
        cell.font = bold_font
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.border = border
        cell.alignment = center_align

    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 6
    for col in range(3, col_idx):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Freeze panes
    ws.freeze_panes = 'C3'

    # Save workbook
    wb.save(output_file)
    print(f"Excel spreadsheet created: {output_file}")

def main():
    script_dir = Path(__file__).parent.parent  # Go up one level from utils/
    ontology_file = script_dir / 'rdf' / 'abr_ontology.ttl'
    output_file = script_dir / 'deliverables' / 'ABR_State_Requirements_Analysis.xlsx'

    print("Parsing ABR ontology...")
    state_info = parse_ontology(ontology_file)

    print(f"\nStates with forms analyzed: {len(state_info)}")

    print("\nGenerating Excel spreadsheet...")
    create_excel_spreadsheet(state_info, output_file)

    # Calculate and display summary statistics
    print("\n" + "="*60)
    print("SUMMARY STATISTICS")
    print("="*60)

    # Most commonly required information
    info_counts = defaultdict(int)
    for state_data in state_info.values():
        for info_type in state_data:
            info_counts[info_type] += 1

    print("\nMost commonly required information types:")
    sorted_counts = sorted(info_counts.items(), key=lambda x: x[1], reverse=True)
    for info_type, count in sorted_counts[:10]:
        pct = (count / len(state_info)) * 100
        label = INFO_LABELS.get(info_type, info_type)
        print(f"  {label:40} {count:2} states ({pct:5.1f}%)")

    print(f"\nSpreadsheet ready for presentation to Rock the Vote!")
    print(f"Location: {output_file}")

if __name__ == '__main__':
    main()
